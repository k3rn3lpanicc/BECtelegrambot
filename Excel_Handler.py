import sqlite3

import openpyxl
import re
from openpyxl.styles.alignment import Alignment
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
import Userhandle

def readTable(filename,columns):
    wb_obj = openpyxl.load_workbook(filename) #get a excel object
    sheet_obj = wb_obj.active #get it's main sheet
    if(columns==[]): #if column addresses were empty then what we want to read ?
        return []
    data = [] # the thing we want to return , each row is a dict (json format)
    col_names =[] #it stores the alphabatic part of each column (AD12 => AD)
    col = [] #it stores the name of column : id , name , etc
    for column in columns:
        col_names.append("")
        for i in range(len(column)):
            if(not str.isdigit(column[i])): #we add the alphabatic part till it is alphabatic
                col_names[-1]+=column[i]
        if(sheet_obj[column].value == None):
            return []
        col.append(sheet_obj[column].value) #column names
    i=int(re.search(r'\d+', columns[0]).group()) + 1 #get the int part of the column address , A12 => 12
    while(True):
        if(sheet_obj[col_names[0]+str(i)].value == None): #when we reached to the end of the table
            break
        dat = dict()
        for j in range(len(col)):
            dat[col[j]] = sheet_obj[col_names[j]+str(i)].value
        data.append(dat)
        i+=1
    return data

def writeTable(columnnames , columns_addr , rows , filename):
    column_fill = PatternFill(start_color='54b4d3', end_color='54b4d3', fill_type='solid')
    greenFill = PatternFill(start_color='d0f0c0', end_color='d0f0c0', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))
    thick_border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'),
                         bottom=Side(style='thick'))

    wb = openpyxl.Workbook()
    sheet = wb.active
    if(len(columnnames)!=len(columns_addr)):
        return False
    # Set the column names in the excel file
    for i in range(len(columnnames)):
        sheet[columns_addr[i]].value = columnnames[i]

    col_addr = [] #the alphabatic part of column addresses
    for i in range(len(columns_addr)):
        sheet[columns_addr[i]].alignment = Alignment(horizontal="center", vertical="center")
        sheet[columns_addr[i]].fill = column_fill
        sheet[columns_addr[i]].border = thick_border

        adr = ""
        for j in range(len(columns_addr[i])):
            if(not str.isdigit(columns_addr[i][j])):
                adr+=columns_addr[i][j]
        sheet.column_dimensions[adr].width = 32
        col_addr.append(adr)
    i = int(re.search(r'\d+', columns_addr[0]).group()) + 1  # get the int part of the column address , A12 => 12
    #now it's time to put the rows in the excel file
    sheet.row_dimensions[1].height = 30

    for row in rows:
        for j in range(len(col_addr)):
            sheet[col_addr[j]+str(i)].value = row[j]
            sheet[col_addr[j]+str(i)].alignment = Alignment(horizontal="center",vertical="center")
            sheet[col_addr[j] + str(i)].fill = greenFill
            sheet[col_addr[j] + str(i)].border = thin_border
            sheet.row_dimensions[i].height = 20
        i+=1
    wb.save(filename) #save it to excel file

def nnm(n):
	return chr(n-1+ord('A'))
def num_to_name(n):
	ans = []
	ret = ""
	while(n>26):
		if(n%26!=0):
			ans.append(n%26)
		else:
			ans.append(26)
		n=n//26
	if(n%26!=0):
		ans.append(n%26)
	else:
		ans.append(26)
	ans.reverse()
	for i in ans:
		ret+=nnm(i)
	return ret

def save_events(file_name):
    column_fill = PatternFill(start_color='54b4d3', end_color='54b4d3', fill_type='solid')
    greenFill = PatternFill(start_color='d0f0c0', end_color='d0f0c0', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))
    thick_border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'),
                          bottom=Side(style='thick'))

    events = Userhandle.get_all_events()
    if(events == False):
        return "No_event"
    columnnames = []
    columnnames2 = []
    columns_addr = []
    cnt = 1
    for event in events:
        columns_addr.append(num_to_name(cnt)+"1")
        cnt+=1
        columnnames.append(event['event_type']+"-"+event['event_name'])
        columnnames2.append(event['id'])
    wb = openpyxl.Workbook()
    sheet = wb.active
    for i in range(len(columnnames)):
        sheet[columns_addr[i]].value = columnnames[i]

    col_addr = []  # the alphabatic part of column addresses
    for i in range(len(columns_addr)):
        sheet[columns_addr[i]].alignment = Alignment(horizontal="center", vertical="center")
        sheet[columns_addr[i]].fill = column_fill
        sheet[columns_addr[i]].border = thick_border
        adr = ""
        for j in range(len(columns_addr[i])):
            if (not str.isdigit(columns_addr[i][j])):
                adr += columns_addr[i][j]
        sheet.column_dimensions[adr].width = 64

        col_addr.append(adr)
    i2 = int(re.search(r'\d+', columns_addr[0]).group()) + 1  # get the int part of the column address , A12 => 12
    sheet.row_dimensions[1].height = 35
    for i in range(len(columns_addr)):
        rows = [events[i]['sign_ups'].split(',')[k] for k in range(len(events[i]['sign_ups'].split(','))) if (not events[i]['sign_ups'].split(',')[k] in ['' , None])]
        #print(rows)
        for j in range(len(rows)):
            sheet.row_dimensions[(j + 2)].height = 28
            print(rows[j])
            user = None
            hamrah = ""
            melliCode = ""
            if('$' in rows[j]):
                user = Userhandle.get_user_data_by_id(rows[j].split('$')[0])
                hamrah = rows[j].split('$')[1]
                melliCode = rows[j].split('$')[2]
            else:
                user = Userhandle.get_user_data_by_id(rows[j])
            if(user!=False):
                txt = str(user['name']) +"("+str(user['id'])+")-"+str(user['melli_code'] if user['melli_code']!=None else "")
                if(hamrah!=""):
                    txt+=" همراه : " + hamrah+"("+melliCode+")"
                sheet[col_addr[i]+str(j+2)].value = txt #str(user['name']) +"("+str(user['id'])+")-"+str(user['melli_code'] if user['melli_code']!=None else "")
                sheet[col_addr[i] + str(j + 2)].border = thin_border
                sheet[col_addr[i] + str(j + 2)].fill = greenFill
                sheet[col_addr[i] + str(j + 2)].alignment = Alignment(horizontal="center", vertical="center")

    wb.save(file_name)
    return "done"

def save_events2(file_name):
    column_fill = PatternFill(start_color='54b4d3', end_color='54b4d3', fill_type='solid')
    greenFill = PatternFill(start_color='d0f0c0', end_color='d0f0c0', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))
    thick_border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'),
                          bottom=Side(style='thick'))

    events = Userhandle.get_all_events2()
    if(events == False):
        return "No_event"
    columnnames = []
    columnnames2 = []
    columns_addr = []
    cnt = 1
    for event in events:
        columns_addr.append(num_to_name(cnt)+"1")
        cnt+=1
        columnnames.append(event['name'])
        columnnames2.append(event['id'])
    wb = openpyxl.Workbook()
    sheet = wb.active
    for i in range(len(columnnames)):
        sheet[columns_addr[i]].value = columnnames[i]

    col_addr = []  # the alphabatic part of column addresses
    for i in range(len(columns_addr)):
        sheet[columns_addr[i]].alignment = Alignment(horizontal="center", vertical="center")
        sheet[columns_addr[i]].fill = column_fill
        sheet[columns_addr[i]].border = thick_border
        adr = ""
        for j in range(len(columns_addr[i])):
            if (not str.isdigit(columns_addr[i][j])):
                adr += columns_addr[i][j]
        sheet.column_dimensions[adr].width = 32

        col_addr.append(adr)
    i2 = int(re.search(r'\d+', columns_addr[0]).group()) + 1  # get the int part of the column address , A12 => 12
    sheet.row_dimensions[1].height = 35
    for i in range(len(columns_addr)):
        print(events[i]['signedtcodes'])
        rows = []

        if(events[i]['signedtcodes']!=None):
            rows = [(events[i]['signedtcodes'].split(',')[k].split(':')[1]+"-"+events[i]['signedtcodes'].split(',')[k].split(':')[2]) for k in range(len(events[i]['signedtcodes'].split(','))) if (not events[i]['signedtcodes'].split(',')[k] in ['' , None])]
        for j in range(len(rows)):
            sheet.row_dimensions[(j + 2)].height = 28
            sheet[col_addr[i]+str(j+2)].value = rows[j]#str(user['name']) +"("+str(user['id'])+")-"+str(user['melli_code'] if user['melli_code']!=None else "")
            sheet[col_addr[i] + str(j + 2)].border = thin_border
            sheet[col_addr[i] + str(j + 2)].fill = greenFill
            sheet[col_addr[i] + str(j + 2)].alignment = Alignment(horizontal="center", vertical="center")

    wb.save(file_name)
    return "done"
